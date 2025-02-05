import pandas as pd
import glob
import os
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def combine_excel_files():
    # Get list of Excel files matching pattern
    print("\nSearching for Excel files with pattern 'signal_mt_stop_*.xlsx'...")
    excel_files = glob.glob('data/signal_mt_stop_*.xlsx')
    
    if not excel_files:
        print("No matching Excel files found in data directory")
        return
    
    print(f"\nFound {len(excel_files)} files to process")
    
    # Read and combine all Excel files
    df_combined = pd.DataFrame()
    for idx, file in enumerate(excel_files, 1):
        print(f"\nProcessing file {idx}/{len(excel_files)}: {os.path.basename(file)}")
        df = pd.read_excel(file)
        print(f"- Rows in file: {len(df)}")
        df_combined = pd.concat([df_combined, df], ignore_index=True)
        print(f"- Total combined rows so far: {len(df_combined)}")
    
    print("\nSorting data by DATE column...")
    if 'DATE' in df_combined.columns:
        df_combined.sort_values('DATE', inplace=True)
    
    # Export combined data
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"data/all_signal_mt_stop_{now}.xlsx"
    print(f"\nExporting combined data to: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        print("- Writing data to Excel...")
        df_combined.to_excel(writer, sheet_name='signal_mt_stop', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['signal_mt_stop']
        
        print("- Adjusting column widths...")
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    print(f"  Warning: Error processing cell {cell}: {e}")
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        
        print("- Adding borders and formatting...")
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.fill = PatternFill(start_color='FFFFFF', 
                                      end_color='FFFFFF', 
                                      fill_type='solid')
        
        print("- Adding auto-filter...")
        worksheet.auto_filter.ref = worksheet.dimensions
    
    print("\nProcess completed successfully!")
    print(f"Total files processed: {len(excel_files)}")
    print(f"Total rows in combined file: {len(df_combined)}")
    print(f"Output file: {output_file}")

if __name__ == "__main__":
    combine_excel_files()